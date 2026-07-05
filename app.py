"""
US Market RSI Screener — Streamlit Web App
══════════════════════════════════════════
Three screeners:
  1. RSI Screen     — RSI(9) overbought / oversold, NASDAQ+NYSE+AMEX
  2. Green Line     — EMA(250) + IBD-RS > 80 + Slow Stochastic(5,1) < 20 turning up
  3. S&P 500 Movers — Top 3 gainers / losers by daily % change

Calculator: RSI & Std Dev / Z-Score Excel builder (Yahoo Finance + Twelve Data)
Pattern Guide: Pattern reference with descriptions
"""

import io
import time
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import requests
import streamlit as st
from plotly.subplots import make_subplots

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule, ColorScaleRule, DataBarRule
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference, Series
    HAS_XL = True
except ImportError:
    HAS_XL = False


# ══════════════════════════════════════════════════════════════════
#  PAGE CONFIG
# ══════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="US Market Screener",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    [data-testid="stSidebar"] { background-color: #1a2235; }
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span { color: #a8b4c4 !important; font-size: 0.88rem; }
    [data-testid="stSidebar"] input { color: #c9d1d9 !important; background-color: #253047 !important; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] {
        background-color: #1e2736; border-radius: 6px 6px 0 0;
        padding: 8px 20px; color: #8b949e;
    }
    .stTabs [aria-selected="true"] {
        background-color: #253047 !important; color: #c9d1d9 !important;
    }
    h1, h2, h3, p, span, label { color: #c9d1d9; }
    div[data-testid="stDataFrame"] { border-radius: 8px; }
    .stButton > button { border-radius: 6px; font-weight: 600; }
    [data-testid="stCaptionContainer"] p { color: #6e7f96 !important; }
    .metric-box {
        background: #1e2736; border-radius: 8px; padding: 14px 18px;
        border-left: 4px solid; margin-bottom: 8px;
    }
    .metric-green { border-color: #3fb950; }
    .metric-red   { border-color: #f85149; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
#  INDICATORS
# ══════════════════════════════════════════════════════════════════

def wilder_rsi(closes: np.ndarray, period: int) -> float:
    need = period * 3
    if len(closes) < need:
        return float("nan")
    arr = closes[-need:].astype(float)
    d = np.diff(arr)
    g, l = np.maximum(d, 0.0), np.maximum(-d, 0.0)
    ag, al = g[:period].mean(), l[:period].mean()
    for i in range(period, len(d)):
        ag = (ag * (period - 1) + g[i]) / period
        al = (al * (period - 1) + l[i]) / period
    return 100.0 if al == 0 else 100.0 - 100.0 / (1.0 + ag / al)


def rsi_series(closes: np.ndarray, period: int) -> np.ndarray:
    n, out = len(closes), np.full(len(closes), np.nan)
    if n < period + 1:
        return out
    c = closes.astype(float)
    d = np.diff(c)
    g, l = np.maximum(d, 0.0), np.maximum(-d, 0.0)
    ag, al = g[:period].mean(), l[:period].mean()

    def _r(ag, al):
        return 100.0 if al == 0 else 100.0 - 100.0 / (1.0 + ag / al)

    out[period] = _r(ag, al)
    for i in range(period, len(d)):
        ag = (ag * (period - 1) + g[i]) / period
        al = (al * (period - 1) + l[i]) / period
        out[i + 1] = _r(ag, al)
    return out


def slow_stochastic(hist: pd.DataFrame, k_period: int = 5) -> tuple[float, float]:
    """
    Slow Stochastic %K(k_period, 1).
    Returns (current_k, previous_k). Both NaN if insufficient data.
    Smoothing=1 means raw %K — no additional SMA applied.
    """
    if len(hist) < k_period + 1:
        return float("nan"), float("nan")
    hi = hist["High"].values.astype(float)
    lo = hist["Low"].values.astype(float)
    cl = hist["Close"].values.astype(float)

    def _k(idx):
        window_lo = lo[idx - k_period + 1: idx + 1]
        window_hi = hi[idx - k_period + 1: idx + 1]
        rng = window_hi.max() - window_lo.min()
        if rng == 0:
            return 50.0
        return (cl[idx] - window_lo.min()) / rng * 100.0

    n = len(cl)
    curr = _k(n - 1)
    prev = _k(n - 2)
    return curr, prev


def ema_series(closes: np.ndarray, period: int) -> np.ndarray:
    """Exponential moving average series."""
    out = np.full(len(closes), np.nan)
    if len(closes) < period:
        return out
    k = 2.0 / (period + 1)
    out[period - 1] = closes[:period].mean()
    for i in range(period, len(closes)):
        out[i] = closes[i] * k + out[i - 1] * (1 - k)
    return out


# ══════════════════════════════════════════════════════════════════
#  PATTERN DETECTION
# ══════════════════════════════════════════════════════════════════

def _peaks(a, dist=5):
    return [i for i in range(dist, len(a) - dist)
            if a[i] == a[i - dist: i + dist + 1].max()]

def _troughs(a, dist=5):
    return [i for i in range(dist, len(a) - dist)
            if a[i] == a[i - dist: i + dist + 1].min()]


def detect_patterns(hist: pd.DataFrame) -> str:
    if len(hist) < 30:
        return "Insufficient data"
    c  = hist["Close"].values.astype(float)
    h  = hist["High"].values.astype(float)
    lo = hist["Low"].values.astype(float)
    v  = hist["Volume"].values.astype(float)
    n  = len(c)
    avg_v20 = v[-20:].mean()
    found = []

    if c[-1] > h[-21:-1].max() and v[-1] > 1.5 * avg_v20: found.append("Breakout ↑")
    if c[-1] < lo[-21:-1].min() and v[-1] > 1.5 * avg_v20: found.append("Breakdown ↓")

    pk = _peaks(h)
    if len(pk) >= 2:
        p1, p2 = pk[-2], pk[-1]
        if p2 > p1 and abs(h[p1] - h[p2]) / h[p1] < 0.03:
            if c[p1:p2 + 1].min() < min(h[p1], h[p2]) * 0.97:
                found.append("Double Top")

    tr = _troughs(lo)
    if len(tr) >= 2:
        t1, t2 = tr[-2], tr[-1]
        if t2 > t1 and abs(lo[t1] - lo[t2]) / lo[t1] < 0.03:
            if c[t1:t2 + 1].max() > max(lo[t1], lo[t2]) * 1.03:
                found.append("Double Bottom")

    if n >= 16:
        pole = (c[-11] - c[-16]) / c[-16]
        flag = (c[-1] - c[-11]) / c[-11]
        rng  = (c[-6:].max() - c[-6:].min()) / c[-11]
        if pole > 0.05 and -0.03 < flag < 0.01 and rng < 0.04: found.append("Bull Flag")
        if pole < -0.05 and -0.01 < flag < 0.03 and rng < 0.04: found.append("Bear Flag")

    if n >= 20:
        x = np.arange(20, dtype=float)
        rh, rl = h[-20:], lo[-20:]
        if np.std(rh[-10:]) / rh[-10:].mean() < 0.015 and np.polyfit(x, rl, 1)[0] > 0:
            found.append("Asc. Triangle")
        if np.std(rl[-10:]) / rl[-10:].mean() < 0.015 and np.polyfit(x, rh, 1)[0] < 0:
            found.append("Desc. Triangle")

    if n >= 40:
        lrim = c[:10].max(); rrim = c[-10:].max(); cup = c[10:n - 10].min()
        rim = min(lrim, rrim)
        if (rim - cup) / rim > 0.12 and abs(lrim - rrim) / lrim < 0.05:
            found.append("Cup & Handle")

    return ", ".join(found) if found else "—"


# ══════════════════════════════════════════════════════════════════
#  PATTERN DESCRIPTIONS
# ══════════════════════════════════════════════════════════════════

PATTERN_DESCRIPTIONS = {
    "Breakout ↑": (
        "Momentum breakout above resistance",
        "Price cleared its 20-bar high today on volume 1.5× the norm. "
        "Buyers stepped in with conviction at a level that had previously "
        "acted as a ceiling. Watch for follow-through in the next 1–2 sessions — "
        "a failure to hold the breakout level is a red flag.",
    ),
    "Breakdown ↓": (
        "Momentum breakdown below support",
        "Price broke below its 20-bar low on elevated volume. Sellers overwhelmed "
        "a level that had held as a floor. Risk of further downside is elevated — "
        "prior support often becomes new resistance on any bounce.",
    ),
    "Double Top": (
        "Bearish reversal — two failed attempts at the same high",
        "Two peaks formed at roughly the same price, separated by a valley (the neckline). "
        "Sellers defended that level twice. A close below the neckline is the trigger — "
        "downside target is the pattern depth subtracted from it.",
    ),
    "Double Bottom": (
        "Bullish reversal — two failed attempts at the same low",
        "Two troughs at similar levels with a peak (neckline) between them. Buyers "
        "absorbed selling twice. A close above the neckline triggers the pattern — "
        "upside target is the depth added to the neckline.",
    ),
    "Bull Flag": (
        "Bullish continuation — tight consolidation after a surge",
        "A sharp upward move (pole) followed by 4–5 bars of sideways drift on lower "
        "volume. The market is catching its breath. Breakout above the flag targets "
        "pole height added to the breakout point.",
    ),
    "Bear Flag": (
        "Bearish continuation — tight consolidation after a drop",
        "A sharp downward move followed by a brief, tight bounce on low volume. "
        "Sellers are pausing, not capitulating. Breakdown below the flag targets "
        "the pole depth subtracted from the breakdown.",
    ),
    "Asc. Triangle": (
        "Bullish bias — flat resistance, rising lows",
        "Resistance holds at a fixed price but each pullback finds support higher — "
        "buyers are getting more aggressive. Breakout targets the triangle height "
        "added to the breakout.",
    ),
    "Desc. Triangle": (
        "Bearish bias — flat support, falling highs",
        "Support holds but sellers push highs lower on each rally. Breakdown targets "
        "the triangle height subtracted from the breakdown.",
    ),
    "Cup & Handle": (
        "Bullish continuation — U-shaped base with a small pullback",
        "A rounded U-shaped recovery back to prior highs (cup) followed by a small "
        "orderly pullback (handle). Breakout above the rim targets the cup depth "
        "added to the breakout. One of the most reliable intermediate-term bullish setups.",
    ),
}


# ══════════════════════════════════════════════════════════════════
#  UNIVERSE FETCH  (shared by RSI and Green Line screens)
# ══════════════════════════════════════════════════════════════════

_REQ_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept": "application/json",
}


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_us_universe(min_price: float) -> list[str]:
    """Pull NASDAQ + NYSE + AMEX, pre-filter by price. Cached 1hr."""
    seen: set[str] = set()
    out:  list[str] = []
    for exchange in ("nasdaq", "nyse", "amex"):
        url = (
            "https://api.nasdaq.com/api/screener/stocks"
            f"?tableonly=true&limit=5000&offset=0&exchange={exchange}"
        )
        try:
            rows = (
                requests.get(url, headers=_REQ_HEADERS, timeout=25)
                        .json()["data"]["table"]["rows"]
            )
        except Exception:
            continue
        for r in rows:
            sym = r.get("symbol", "").strip()
            if not sym or "/" in sym or "^" in sym or len(sym) > 5:
                continue
            if sym in seen:
                continue
            try:
                p = float(r.get("lastsale", "$0").replace("$", "").replace(",", ""))
            except ValueError:
                continue
            if p >= min_price * 0.90:
                seen.add(sym)
                out.append(sym)
    return out


# ══════════════════════════════════════════════════════════════════
#  SCREENER 1 — RSI(9)
# ══════════════════════════════════════════════════════════════════

def screen_rsi(
    tickers, min_price, min_avg_vol, min_curr_vol,
    rsi9_ob, rsi9_os, progress_cb=None,
) -> tuple[list[dict], list[dict], dict]:
    import yfinance as yf
    ob, os_, cache = [], [], {}
    total, BATCH = len(tickers), 40

    for i in range(0, total, BATCH):
        batch = tickers[i: i + BATCH]
        if progress_cb:
            progress_cb(int(i / total * 50),
                        f"Downloading batch {i // BATCH + 1} / {-(-total // BATCH)}")
        try:
            raw = yf.download(tickers=batch, period="4mo", interval="1d",
                              auto_adjust=True, progress=False, group_by="ticker")
            if not raw.empty:
                if len(batch) == 1:
                    sub = raw.copy()
                    if isinstance(sub.columns, pd.MultiIndex):
                        sub.columns = sub.columns.droplevel(1)
                    if not sub.empty:
                        cache[batch[0]] = sub
                else:
                    for t in batch:
                        try:
                            sub = raw[t].dropna(how="all")
                            if not sub.empty:
                                cache[t] = sub
                        except Exception:
                            pass
        except Exception:
            pass
        time.sleep(0.25)

    n_h = len(cache)
    for idx, (t, hist) in enumerate(cache.items()):
        if progress_cb:
            progress_cb(50 + int(idx / max(n_h, 1) * 50),
                        f"Screening {t}  ({idx + 1} / {n_h})")
        hist = hist.dropna(subset=["Close", "Volume"])
        if len(hist) < 20:
            continue
        c = hist["Close"].values.astype(float)
        v = hist["Volume"].values.astype(float)
        price, curr_vol, avg_vol = c[-1], float(v[-1]), float(v[-20:].mean())
        if price < min_price or avg_vol < min_avg_vol or curr_vol < min_curr_vol:
            continue
        r9 = wilder_rsi(c, 9)
        if np.isnan(r9) or (rsi9_os <= r9 <= rsi9_ob):
            continue
        pattern = detect_patterns(hist)
        row = dict(Ticker=t, Price=f"${price:.2f}",
                   CurrVol=f"{curr_vol:,.0f}", AvgVol=f"{avg_vol:,.0f}",
                   RSI9=f"{r9:.1f}", Pattern=pattern, _rsi9=r9)
        if r9 > rsi9_ob: ob.append(row)
        if r9 < rsi9_os: os_.append(row)

    ob.sort(key=lambda r: r["_rsi9"], reverse=True)
    os_.sort(key=lambda r: r["_rsi9"])
    return ob, os_, cache


# ══════════════════════════════════════════════════════════════════
#  SCREENER 2 — GREEN LINE
#  Rules: Close > EMA(250)  AND  IBD-RS > 80  AND
#         Slow Stochastic(5,1) < 20 AND today > yesterday
# ══════════════════════════════════════════════════════════════════

def screen_green_line(
    tickers, min_price, min_avg_vol, rs_threshold,
    progress_cb=None,
) -> tuple[list[dict], dict]:
    """
    Requires 15 months of daily OHLCV per ticker:
      - EMA(250) needs 250 trading days
      - IBD-RS needs 252 trading days (12-month return)
      - Stochastic(5,1) needs 5 days

    IBD-RS approximation: 12-month price return percentile within
    the scan universe. Not identical to SCTR — peer ranking is
    the unresolvable gap without a StockCharts API subscription.

    Slow Stochastic trigger: loose — %K < 20 AND today > yesterday
    (entry while still in green zone, matching the chart examples).
    """
    import yfinance as yf

    results = []
    cache   = {}
    total, BATCH = len(tickers), 25   # smaller batches — 15mo is heavy

    # ── Download 15mo daily bars ────────────────────────────────
    for i in range(0, total, BATCH):
        batch = tickers[i: i + BATCH]
        if progress_cb:
            progress_cb(
                int(i / total * 60),
                f"Downloading batch {i // BATCH + 1} / {-(-total // BATCH)}  "
                f"(15mo data — this scan runs slower)"
            )
        try:
            raw = yf.download(tickers=batch, period="15mo", interval="1d",
                              auto_adjust=True, progress=False, group_by="ticker")
            if not raw.empty:
                if len(batch) == 1:
                    sub = raw.copy()
                    if isinstance(sub.columns, pd.MultiIndex):
                        sub.columns = sub.columns.droplevel(1)
                    if not sub.empty:
                        cache[batch[0]] = sub
                else:
                    for t in batch:
                        try:
                            sub = raw[t].dropna(how="all")
                            if not sub.empty:
                                cache[t] = sub
                        except Exception:
                            pass
        except Exception:
            pass
        time.sleep(0.3)

    # ── Compute 12-month returns for RS percentile ───────────────
    returns_12m: dict[str, float] = {}
    for t, hist in cache.items():
        hist = hist.dropna(subset=["Close"])
        if len(hist) < 252:
            continue
        price_now  = float(hist["Close"].values[-1])
        price_year = float(hist["Close"].values[-252])
        if price_year > 0:
            returns_12m[t] = (price_now - price_year) / price_year

    if not returns_12m:
        return [], cache

    # Rank into percentiles within this scan universe
    sorted_tickers = sorted(returns_12m, key=returns_12m.get)
    rs_rank = {
        t: (i / max(len(sorted_tickers) - 1, 1)) * 100
        for i, t in enumerate(sorted_tickers)
    }

    # ── Screen each ticker ───────────────────────────────────────
    n_h = len(cache)
    for idx, (t, hist) in enumerate(cache.items()):
        if progress_cb:
            progress_cb(
                60 + int(idx / max(n_h, 1) * 40),
                f"Screening {t}  ({idx + 1} / {n_h})"
            )
        hist = hist.dropna(subset=["Close", "High", "Low", "Volume"])
        if len(hist) < 252:
            continue

        c   = hist["Close"].values.astype(float)
        v   = hist["Volume"].values.astype(float)

        price   = c[-1]
        avg_vol = float(v[-20:].mean())

        if price < min_price or avg_vol < min_avg_vol:
            continue

        # Rule 1 — above EMA(250)
        ema250 = ema_series(c, 250)
        if np.isnan(ema250[-1]) or price <= ema250[-1]:
            continue

        # Rule 2 — IBD-RS percentile > threshold
        rs = rs_rank.get(t, 0.0)
        if rs < rs_threshold:
            continue

        # Rule 3 — Slow Stochastic(5,1) < 20 AND turning up
        k_curr, k_prev = slow_stochastic(hist, k_period=5)
        if np.isnan(k_curr) or np.isnan(k_prev):
            continue
        if k_curr >= 20 or k_curr <= k_prev:
            continue

        pct_above_ema = (price - ema250[-1]) / ema250[-1] * 100
        results.append(dict(
            Ticker        = t,
            Price         = f"${price:.2f}",
            AvgVol        = f"{avg_vol:,.0f}",
            EMA250        = f"${ema250[-1]:.2f}",
            PctAboveEMA   = f"{pct_above_ema:.1f}%",
            RS            = f"{rs:.0f}",
            Stoch         = f"{k_curr:.1f}",
            _rs           = rs,
        ))

    results.sort(key=lambda r: r["_rs"], reverse=True)
    return results, cache


# ══════════════════════════════════════════════════════════════════
#  SCREENER 3 — S&P 500 MOVERS
#  Top 3 gainers + top 3 losers by daily % change
#  Universe: S&P 500 components from Wikipedia (stays current)
# ══════════════════════════════════════════════════════════════════

@st.cache_data(ttl=3600, show_spinner=False)
def fetch_sp500_tickers() -> list[str]:
    """Pull S&P 500 constituent list from Wikipedia. Cached 1hr."""
    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    try:
        tables = pd.read_html(url)
        tickers = tables[0]["Symbol"].tolist()
        # Wikipedia uses dots (BRK.B) — yfinance uses dashes (BRK-B)
        return [t.replace(".", "-") for t in tickers]
    except Exception:
        return []


def screen_sp500_movers(progress_cb=None) -> list[dict]:
    """
    Pulls 2 days of daily OHLCV for all S&P 500 components,
    computes daily % change, returns top 3 + bottom 3.
    """
    import yfinance as yf

    tickers = fetch_sp500_tickers()
    if not tickers:
        return []

    if progress_cb:
        progress_cb(10, f"Fetching {len(tickers)} S&P 500 components…")

    BATCH = 50
    results = []

    for i in range(0, len(tickers), BATCH):
        batch = tickers[i: i + BATCH]
        if progress_cb:
            progress_cb(
                10 + int(i / len(tickers) * 80),
                f"Downloading batch {i // BATCH + 1} / {-(-len(tickers) // BATCH)}"
            )
        try:
            raw = yf.download(tickers=batch, period="5d", interval="1d",
                              auto_adjust=True, progress=False, group_by="ticker")
            if raw.empty:
                continue

            for t in batch:
                try:
                    if len(batch) == 1:
                        sub = raw.copy()
                        if isinstance(sub.columns, pd.MultiIndex):
                            sub.columns = sub.columns.droplevel(1)
                    else:
                        sub = raw[t].dropna(how="all")
                    if sub.empty or len(sub) < 2:
                        continue
                    closes = sub["Close"].dropna().values.astype(float)
                    if len(closes) < 2:
                        continue
                    prev_close = closes[-2]
                    curr_close = closes[-1]
                    if prev_close <= 0:
                        continue
                    pct = (curr_close - prev_close) / prev_close * 100
                    vol = float(sub["Volume"].values[-1]) if "Volume" in sub.columns else 0
                    results.append(dict(
                        Ticker  = t,
                        Price   = f"${curr_close:.2f}",
                        Change  = f"{pct:+.2f}%",
                        Volume  = f"{vol:,.0f}",
                        _pct    = pct,
                    ))
                except Exception:
                    pass
        except Exception:
            pass
        time.sleep(0.2)

    if not results:
        return []

    results.sort(key=lambda r: r["_pct"], reverse=True)
    gainers = results[:3]
    losers  = results[-3:][::-1]   # worst first

    # Tag each row
    for r in gainers: r["Side"] = "Gainer"
    for r in losers:  r["Side"] = "Loser"

    return gainers + losers


# ══════════════════════════════════════════════════════════════════
#  PLOTLY CHART
# ══════════════════════════════════════════════════════════════════

def make_chart(ticker: str, hist: pd.DataFrame, pattern: str,
               ob_thr: float = 80, os_thr: float = 20) -> go.Figure:
    c  = hist["Close"].values.astype(float)
    v  = hist["Volume"].values.astype(float)
    dt = hist.index
    rs = rsi_series(c, 9)

    bar_colors = ["#3fb950" if i == 0 or c[i] >= c[i - 1] else "#f85149"
                  for i in range(len(c))]

    fig = make_subplots(
        rows=3, cols=1, shared_xaxes=True,
        row_heights=[0.58, 0.16, 0.26], vertical_spacing=0.03,
        subplot_titles=[f"<b>{ticker}</b>  —  {pattern}", "Volume", "RSI(9)"],
    )
    fig.add_trace(go.Scatter(
        x=dt, y=c, fill="tozeroy", fillcolor="rgba(63,185,80,0.08)",
        line=dict(color="#3fb950", width=1.5), name="Close",
        hovertemplate="$%{y:.2f}<extra></extra>",
    ), row=1, col=1)
    fig.add_trace(go.Bar(
        x=dt, y=v, marker_color=bar_colors, opacity=0.6, name="Volume",
        hovertemplate="%{y:,.0f}<extra></extra>",
    ), row=2, col=1)
    fig.add_hrect(y0=ob_thr, y1=100, fillcolor="rgba(248,81,73,0.08)",
                  line_width=0, row=3, col=1)
    fig.add_hrect(y0=0, y1=os_thr, fillcolor="rgba(63,185,80,0.08)",
                  line_width=0, row=3, col=1)
    fig.add_hline(y=ob_thr, line_dash="dash", line_color="#f85149",
                  line_width=1, opacity=0.7, row=3, col=1)
    fig.add_hline(y=os_thr, line_dash="dash", line_color="#3fb950",
                  line_width=1, opacity=0.7, row=3, col=1)
    fig.add_trace(go.Scatter(
        x=dt, y=rs, line=dict(color="#f0b429", width=1.3), name="RSI(9)",
        hovertemplate="RSI %{y:.1f}<extra></extra>",
    ), row=3, col=1)
    rsi_last = rs[~np.isnan(rs)][-1] if np.any(~np.isnan(rs)) else float("nan")
    fig.update_layout(
        plot_bgcolor="#161b22", paper_bgcolor="#0d1117",
        font=dict(color="#e6edf3", size=11), height=560,
        showlegend=False, margin=dict(l=10, r=10, t=40, b=10),
        hovermode="x unified", hoverlabel=dict(bgcolor="#21262d", font_color="#e6edf3"),
    )
    fig.update_xaxes(gridcolor="#21262d", showgrid=True,
                     zeroline=False, rangeslider_visible=False)
    fig.update_yaxes(gridcolor="#21262d", showgrid=True, zeroline=False)
    fig.update_yaxes(tickprefix="$", row=1, col=1)
    fig.update_yaxes(range=[0, 100], tickvals=[0, os_thr, 50, ob_thr, 100], row=3, col=1)
    if not np.isnan(rsi_last):
        fig.layout.annotations[0].text = (
            f"<b>{ticker}</b>  RSI {rsi_last:.1f}  —  {pattern}"
        )
    return fig


def make_gl_chart(ticker: str, hist: pd.DataFrame) -> go.Figure:
    """Green Line chart: price + EMA(250) + Stochastic(5,1)."""
    c  = hist["Close"].values.astype(float)
    h  = hist["High"].values.astype(float)
    lo = hist["Low"].values.astype(float)
    v  = hist["Volume"].values.astype(float)
    dt = hist.index
    e250 = ema_series(c, 250)

    # Stochastic series
    k_series = np.full(len(c), np.nan)
    for i in range(4, len(c)):
        rng = h[i-4:i+1].max() - lo[i-4:i+1].min()
        if rng > 0:
            k_series[i] = (c[i] - lo[i-4:i+1].min()) / rng * 100

    bar_colors = ["#3fb950" if i == 0 or c[i] >= c[i-1] else "#f85149"
                  for i in range(len(c))]

    fig = make_subplots(
        rows=3, cols=1, shared_xaxes=True,
        row_heights=[0.55, 0.15, 0.30], vertical_spacing=0.03,
        subplot_titles=[f"<b>{ticker}</b>  —  Price + EMA(250)", "Volume", "Slow Stoch(5,1)"],
    )
    fig.add_trace(go.Scatter(
        x=dt, y=c, line=dict(color="#3fb950", width=1.4),
        name="Close", hovertemplate="$%{y:.2f}<extra></extra>",
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=dt, y=e250, line=dict(color="#22c55e", width=1.8, dash="solid"),
        name="EMA(250)", opacity=0.85, hovertemplate="EMA $%{y:.2f}<extra></extra>",
    ), row=1, col=1)
    fig.add_trace(go.Bar(
        x=dt, y=v, marker_color=bar_colors, opacity=0.55, name="Volume",
        hovertemplate="%{y:,.0f}<extra></extra>",
    ), row=2, col=1)
    fig.add_hrect(y0=0,  y1=20, fillcolor="rgba(63,185,80,0.10)",
                  line_width=0, row=3, col=1)
    fig.add_hrect(y0=80, y1=100, fillcolor="rgba(248,81,73,0.10)",
                  line_width=0, row=3, col=1)
    fig.add_hline(y=20, line_dash="dash", line_color="#3fb950",
                  line_width=1, opacity=0.7, row=3, col=1)
    fig.add_hline(y=80, line_dash="dash", line_color="#f85149",
                  line_width=1, opacity=0.7, row=3, col=1)
    fig.add_trace(go.Scatter(
        x=dt, y=k_series, line=dict(color="#f0b429", width=1.3),
        name="Stoch%K", hovertemplate="%{y:.1f}<extra></extra>",
    ), row=3, col=1)
    fig.update_layout(
        plot_bgcolor="#161b22", paper_bgcolor="#0d1117",
        font=dict(color="#e6edf3", size=11), height=560,
        showlegend=True,
        legend=dict(bgcolor="#1e2736", bordercolor="#253047",
                    font=dict(color="#c9d1d9")),
        margin=dict(l=10, r=10, t=40, b=10),
        hovermode="x unified", hoverlabel=dict(bgcolor="#21262d", font_color="#e6edf3"),
    )
    fig.update_xaxes(gridcolor="#21262d", showgrid=True,
                     zeroline=False, rangeslider_visible=False)
    fig.update_yaxes(gridcolor="#21262d", showgrid=True, zeroline=False)
    fig.update_yaxes(tickprefix="$", row=1, col=1)
    fig.update_yaxes(range=[0, 100], tickvals=[0, 20, 50, 80, 100], row=3, col=1)
    return fig


# ══════════════════════════════════════════════════════════════════
#  EXCEL BUILDERS
# ══════════════════════════════════════════════════════════════════

def _styler():
    def st_cell(c, bold=False, fg="000000", bg=None, align="left",
                sz=10, italic=False, fmt=None):
        c.font      = Font(bold=bold, color=fg, size=sz, italic=italic, name="Arial")
        c.alignment = Alignment(horizontal=align, vertical="center")
        if bg:  c.fill          = PatternFill("solid", fgColor=bg)
        if fmt: c.number_format = fmt
    def bdr(c, color="CCCCCC"):
        s = Side(style="thin", color=color)
        c.border = Border(left=s, right=s, top=s, bottom=s)
    return st_cell, bdr


def build_rsi_excel(df, ticker, period, output) -> None:
    st_cell, bdr = _styler()
    wb = Workbook(); ws = wb.active; ws.title = "RSI"
    P, DS = "$B$2", 5; MAX_R = DS + len(df)
    HDR_BG="1F3864"; HDR_FG="FFFFFF"; INP_BG="FFF2CC"; INP_FG="0000FF"
    OUT_BG="D9E8F5"; WHITE="FFFFFF"; ALT="F2F7FB"; GRAY="888888"
    ws.sheet_view.showGridLines = True
    for col, w in [("A",13),("B",11),("C",12),("D",18),("E",14),
                   ("F",11),("G",11),("H",11),("I",11)]:
        ws.column_dimensions[col].width = w
    for n in range(10, 46):
        ws.column_dimensions[get_column_letter(n)].hidden = True
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"RSI CALCULATOR  ·  {ticker.upper()}  ·  Open · High · Low · Close"
    st_cell(c, bold=True, fg=HDR_FG, bg=HDR_BG, align="center", sz=14)
    ws.row_dimensions[1].height = 32; ws.row_dimensions[2].height = 26
    ws["A2"].value = "RSI Period:"; st_cell(ws["A2"], bold=True, fg=HDR_FG, bg=HDR_BG, sz=10); bdr(ws["A2"],"999999")
    ws["B2"].value = period; st_cell(ws["B2"], bold=True, fg=INP_FG, bg=INP_BG, align="center", fmt="0", sz=11); bdr(ws["B2"],"999999")
    ws["C2"].value = "← Period"; st_cell(ws["C2"], fg=GRAY, bg="F5F5F5", sz=9, italic=True); bdr(ws["C2"],"DDDDDD")
    ws["D2"].value = "Reverse RSI Targets ↓"; st_cell(ws["D2"], bold=True, fg=HDR_FG, bg=HDR_BG, align="center", sz=9); bdr(ws["D2"],"999999")
    ws["E2"].value = "Enter 1-99 per column"; st_cell(ws["E2"], fg=GRAY, bg="F5F5F5", align="center", sz=9, italic=True); bdr(ws["E2"],"DDDDDD")
    for col in ["F","G","H","I"]:
        c = ws[f"{col}2"]; c.value = 70
        st_cell(c, bold=True, fg=INP_FG, bg=INP_BG, align="center", fmt="0.00", sz=11); bdr(c,"FF9900")
    ws.row_dimensions[3].height = 24
    ws["A3"].value = "Required Price:"; st_cell(ws["A3"], bold=True, fg=HDR_FG, bg=HDR_BG, sz=10); bdr(ws["A3"],"999999")
    for col in ["B","C","D","E"]: st_cell(ws[f"{col}3"], bg="F0F4F8"); bdr(ws[f"{col}3"],"DDDDDD")
    SERIES_INFO = [
        ("F2","B","F","M","N",30,34,38,42,6),("G2","C","G","R","S",31,35,39,43,7),
        ("H2","D","H","W","X",32,36,40,44,8),("I2","E","I","AB","AC",33,37,41,45,9),
    ]
    for trsi,pdc,rdc,agdc,aldc,hpc,hrc,hagc,halc,oc in SERIES_INFO:
        for hcol, dcol in [(hpc,pdc),(hrc,rdc),(hagc,agdc),(halc,aldc)]:
            ws.cell(row=3, column=hcol).value = (
                f"=IFERROR(LOOKUP(2,1/({dcol}${DS}:{dcol}${MAX_R}<>\"\"),{dcol}${DS}:{dcol}${MAX_R}),\"\")")
    for trsi,pdc,rdc,agdc,aldc,hpc,hrc,hagc,halc,oc in SERIES_INFO:
        hp=get_column_letter(hpc); hr=get_column_letter(hrc)
        ha=get_column_letter(hagc); hl=get_column_letter(halc)
        c = ws.cell(row=3, column=oc)
        c.value = (f"=IF(OR({trsi}<=0,{trsi}>=100),\"INVALID\","
                   f"IF({ha}3=\"\",\"NO DATA\","
                   f"IF({trsi}>={hr}3,"
                   f"ROUND({hp}3+({P}-1)*({hl}3*{trsi}/(100-{trsi})-{ha}3),2),"
                   f"ROUND({hp}3-({P}-1)*({ha}3*(100-{trsi})/{trsi}-{hl}3),2))))")
        st_cell(c, bold=True, bg=OUT_BG, align="center", fmt="#,##0.00", sz=11); bdr(c,"2E75B6")
    for cn, hdr, bg in [(1,"DATE",HDR_BG),(2,"OPEN",HDR_BG),(3,"HIGH",HDR_BG),
                        (4,"LOW",HDR_BG),(5,"CLOSE",HDR_BG),(6,"RSI (O)","17375E"),
                        (7,"RSI (H)","17375E"),(8,"RSI (L)","17375E"),(9,"RSI (C)","17375E")]:
        c = ws.cell(row=4, column=cn, value=hdr)
        st_cell(c, bold=True, fg=HDR_FG, bg=bg, align="center", sz=10); bdr(c)
    ws.row_dimensions[4].height = 24; ws.freeze_panes = "A5"
    for idx, (date, row_data) in enumerate(df.iterrows()):
        r = DS + idx; abg = ALT if r % 2 == 0 else WHITE
        c = ws.cell(row=r, column=1, value=date.date()); st_cell(c, align="center", sz=10, bg=abg); bdr(c)
        for ci, val in zip([2,3,4,5],[row_data["Open"],row_data["High"],row_data["Low"],row_data["Close"]]):
            cell = ws.cell(row=r, column=ci, value=round(float(val),4))
            st_cell(cell, align="right", fmt="#,##0.00", sz=10, bg=abg); bdr(cell)
    SERIES_DEFS = [("B",10,11,12,13,14,6),("C",15,16,17,18,19,7),
                   ("D",20,21,22,23,24,8),("E",25,26,27,28,29,9)]
    def make_formulas(r, pc, cc, gc, lc, agc, alc, rsi_c):
        CL = get_column_letter
        chg=CL(cc); gain=CL(gc); loss=CL(lc); ag=CL(agc); al=CL(alc)
        gs=f"{gain}${DS+1}"; ls=f"{loss}${DS+1}"
        return {
            cc:  f"=IF({pc}{r}=\"\",\"\",{pc}{r}-{pc}{r-1})",
            gc:  f"=IF({chg}{r}=\"\",\"\",MAX({chg}{r},0))",
            lc:  f"=IF({chg}{r}=\"\",\"\",MAX(-{chg}{r},0))",
            agc: (f"=IF(OR({gain}{r}=\"\",ROW()<{P}+{DS}),\"\","
                  f"IF(ROW()={P}+{DS},AVERAGE(OFFSET({gs},0,0,{P},1)),"
                  f"({ag}{r-1}*({P}-1)+{gain}{r})/{P}))"),
            alc: (f"=IF(OR({loss}{r}=\"\",ROW()<{P}+{DS}),\"\","
                  f"IF(ROW()={P}+{DS},AVERAGE(OFFSET({ls},0,0,{P},1)),"
                  f"({al}{r-1}*({P}-1)+{loss}{r})/{P}))"),
            rsi_c: f"=IF({ag}{r}=\"\",\"\",IF({al}{r}=0,100,ROUND(100-100/(1+{ag}{r}/{al}{r}),2)))",
        }
    for r in range(DS+1, MAX_R+1):
        abg = ALT if r % 2 == 0 else WHITE
        for sd in SERIES_DEFS:
            pc,cc,gc,lc,agc,alc,rsi_c = sd
            for col, fml in make_formulas(r,pc,cc,gc,lc,agc,alc,rsi_c).items():
                ws.cell(row=r, column=col, value=fml)
            rsi_cell = ws.cell(row=r, column=rsi_c)
            st_cell(rsi_cell, bg=abg, align="center", fmt="0.00", sz=10); bdr(rsi_cell)
    RSI_CF = [
        (f"F{DS}>=95","375623","FFFFFF"),(f"F{DS}>=90","548235","FFFFFF"),
        (f"F{DS}>=85","70AD47","FFFFFF"),(f"F{DS}>=80","C6E0B4","375623"),
        (f"F{DS}<=5", "9B0000","FFFFFF"),(f"F{DS}<=10","C00000","FFFFFF"),
        (f"F{DS}<=15","FF6666","FFFFFF"),(f"F{DS}<=20","FFCCCC","9B0000"),
    ]
    for formula, bg_hex, fg_hex in RSI_CF:
        fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
        font = Font(color=fg_hex, name="Arial", size=10)
        dxf  = DifferentialStyle(fill=fill, font=font)
        ws.conditional_formatting.add(f"F{DS}:I{MAX_R}",
                                      Rule(type="expression", dxf=dxf, formula=[formula]))
    wb.save(output)


def build_sd_excel(df, ticker, period, output) -> None:
    st_cell, bdr = _styler()
    wb = Workbook(); ws = wb.active; ws.title = "SD_ZSCORE"
    P, DS = "$B$2", 4; MAX_R = DS + len(df)
    HDR_BG="1F3864"; HDR_FG="FFFFFF"; INP_BG="FFF2CC"
    WHITE="FFFFFF"; ALT="F2F7FB"; GRAY="888888"
    ws.sheet_view.showGridLines = True
    for col, w in [("A",13),("B",11),("C",11),("D",11),("E",11),
                   ("F",10),("G",10),("H",10),("I",10),
                   ("J",9),("K",9),("L",9),("M",9)]:
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = f"STD DEV + Z-SCORE  ·  {ticker.upper()}  ·  Open · High · Low · Close"
    st_cell(c, bold=True, fg=HDR_FG, bg=HDR_BG, align="center", sz=14)
    ws.row_dimensions[1].height = 32; ws.row_dimensions[2].height = 26
    ws["A2"].value = "Rolling Window:"; st_cell(ws["A2"], bold=True, fg=HDR_FG, bg=HDR_BG, sz=10); bdr(ws["A2"],"999999")
    ws["B2"].value = period; st_cell(ws["B2"], bold=True, fg="0000FF", bg="FFF2CC", align="center", fmt="0", sz=11); bdr(ws["B2"],"999999")
    ws.merge_cells("C2:M2"); ws["C2"].value = "← Change window. SD & Z-Score recalc automatically."
    st_cell(ws["C2"], fg=GRAY, bg="F5F5F5", sz=9, italic=True); bdr(ws["C2"],"DDDDDD")
    headers = [(1,"DATE",HDR_BG),(2,"OPEN",HDR_BG),(3,"HIGH",HDR_BG),(4,"LOW",HDR_BG),
               (5,"CLOSE",HDR_BG),(6,"SD (O)","2E75B6"),(7,"SD (H)","2E75B6"),
               (8,"SD (L)","2E75B6"),(9,"SD (C)","2E75B6"),(10,"Z (O)","C00000"),
               (11,"Z (H)","C00000"),(12,"Z (L)","C00000"),(13,"Z (C)","C00000")]
    for cn, hdr, bg in headers:
        c = ws.cell(row=3, column=cn, value=hdr)
        st_cell(c, bold=True, fg=HDR_FG, bg=bg, align="center", sz=10); bdr(c)
    ws.row_dimensions[3].height = 24; ws.freeze_panes = "A4"
    for idx, (date, row_data) in enumerate(df.iterrows()):
        r = DS + idx; abg = ALT if r % 2 == 0 else WHITE
        c = ws.cell(row=r, column=1, value=date.date()); st_cell(c, align="center", sz=10, bg=abg); bdr(c)
        for ci, val in zip([2,3,4,5],[row_data["Open"],row_data["High"],row_data["Low"],row_data["Close"]]):
            cell = ws.cell(row=r, column=ci, value=round(float(val),4))
            st_cell(cell, align="right", fmt="#,##0.00", sz=10, bg=abg); bdr(cell)
    SERIES = [("B",6,10),("C",7,11),("D",8,12),("E",9,13)]
    for r in range(DS, MAX_R+1):
        abg = ALT if r % 2 == 0 else WHITE
        for pc, sd_c, z_c in SERIES:
            sd_letter = get_column_letter(sd_c)
            sd_cell = ws.cell(row=r, column=sd_c)
            sd_cell.value = (f"=IF(ROW()-{DS}+1<{P},\"\","
                             f"ROUND(STDEV(OFFSET({pc}{r},-({P}-1),0,{P},1)),4))")
            st_cell(sd_cell, bg=abg, align="center", fmt="0.0000", sz=10); bdr(sd_cell)
            z_cell = ws.cell(row=r, column=z_c)
            z_cell.value = (f"=IF(ROW()-{DS}+1<{P},\"\","
                            f"IF({sd_letter}{r}=0,0,"
                            f"ROUND(({pc}{r}-AVERAGE(OFFSET({pc}{r},-({P}-1),0,{P},1)))/{sd_letter}{r},2)))")
            st_cell(z_cell, bg=abg, align="center", fmt="0.00", sz=10); bdr(z_cell)
    ws.conditional_formatting.add(f"J{DS}:M{MAX_R}",
        ColorScaleRule(start_type="num", start_value=-3, start_color="2E75B6",
                       mid_type="num", mid_value=0, mid_color="FFFFFF",
                       end_type="num", end_value=3, end_color="C00000"))
    bin_edges = [round(-4 + 0.5*i, 1) for i in range(17)]
    n_bins = len(bin_edges) - 1
    z_cols = {10:"Open",11:"High",12:"Low",13:"Close"}
    series_colors = {10:"888888",11:"548235",12:"C00000",13:"1F3864"}
    z_values = {}
    for col_idx, pname in z_cols.items():
        s = df[pname]
        z = (s - s.rolling(window=period).mean()) / s.rolling(window=period).std()
        z_values[col_idx] = z.dropna().values
    HIST_START_ROW = 1; block_width = 3
    for i, (col_idx, name) in enumerate(z_cols.items()):
        bin_col = 15 + i*block_width; freq_col = bin_col + 1
        bin_letter = get_column_letter(bin_col); freq_letter = get_column_letter(freq_col)
        ws.column_dimensions[bin_letter].width = 9; ws.column_dimensions[freq_letter].width = 8
        ws.merge_cells(start_row=HIST_START_ROW, start_column=bin_col,
                       end_row=HIST_START_ROW, end_column=freq_col)
        tc = ws.cell(row=HIST_START_ROW, column=bin_col,
                     value=f"Z-Score Distribution — {name}")
        st_cell(tc, bold=True, fg="FFFFFF", bg="1F3864", align="center", sz=11)
        hr = HIST_START_ROW + 1
        c1 = ws.cell(row=hr, column=bin_col, value="Z Range")
        c2 = ws.cell(row=hr, column=freq_col, value="Count")
        st_cell(c1, bold=True, fg="FFFFFF", bg=series_colors[col_idx], align="center", sz=9)
        st_cell(c2, bold=True, fg="FFFFFF", bg=series_colors[col_idx], align="center", sz=9)
        zvals = z_values[col_idx]
        for b in range(n_bins):
            lo_b, hi_b = bin_edges[b], bin_edges[b+1]; r = hr + 1 + b
            label = f"{lo_b:+.1f} to {hi_b:+.1f}"
            count = int(((zvals >= lo_b) & (zvals <= hi_b if b == n_bins-1 else zvals < hi_b)).sum())
            lc = ws.cell(row=r, column=bin_col, value=label)
            fc = ws.cell(row=r, column=freq_col, value=count)
            abg = ALT if r % 2 == 0 else WHITE
            st_cell(lc, align="center", sz=9, bg=abg)
            st_cell(fc, align="center", sz=9, bg=abg, bold=True)
        data_range = f"{freq_letter}{hr+1}:{freq_letter}{hr+n_bins}"
        ws.conditional_formatting.add(data_range,
            DataBarRule(start_type="num", start_value=0, end_type="max",
                        color=series_colors[col_idx], showValue=True,
                        minLength=None, maxLength=None))
    wb.save(output)


def download_ohlc(ticker, start, end, interval,
                  source="yahoo", api_key="") -> pd.DataFrame:
    import yfinance as yf

    def _resample(df, iv):
        if iv == "1d": return df
        rule = {"1wk":"W","1mo":"ME","1y":"YE"}[iv]
        return df.resample(rule).agg(
            {"Open":"first","High":"max","Low":"min","Close":"last"}
        ).dropna()

    if source == "twelvedata":
        if not api_key:
            raise ValueError(
                "Twelve Data requires a free API key.\n"
                "Get one at https://twelvedata.com/pricing (no card required)."
            )
        params = dict(symbol=ticker, interval="1day", start_date=start,
                      end_date=end, outputsize=5000, apikey=api_key, format="JSON")
        data = requests.get("https://api.twelvedata.com/time_series",
                            params=params, timeout=25).json()
        if data.get("status") == "error":
            raise ValueError(f"Twelve Data: {data.get('message','unknown error')}")
        if "values" not in data:
            raise ValueError(f"No data for '{ticker}' on Twelve Data.")
        df = pd.DataFrame(data["values"])
        df["datetime"] = pd.to_datetime(df["datetime"])
        df = df.set_index("datetime").sort_index()
        df = df.rename(columns={"open":"Open","high":"High","low":"Low","close":"Close"})
        for col in ["Open","High","Low","Close"]:
            df[col] = df[col].astype(float)
        return _resample(df[["Open","High","Low","Close"]].dropna(), interval)
    else:
        yf_iv = "1mo" if interval == "1y" else interval
        df = yf.download(ticker, start=start, end=end,
                         interval=yf_iv, auto_adjust=False, progress=False)
        if df.empty:
            raise ValueError(f"No data for '{ticker}' on Yahoo Finance.")
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df = df[["Open","High","Low","Close"]].dropna().sort_index()
        return _resample(df, interval)


# ══════════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════════

for key, default in [
    ("rsi_ob",      []),
    ("rsi_os",      []),
    ("rsi_cache",   {}),
    ("rsi_done",    False),
    ("gl_results",  []),
    ("gl_cache",    {}),
    ("gl_done",     False),
    ("sp_results",  []),
    ("sp_done",     False),
]:
    if key not in st.session_state:
        st.session_state[key] = default

RSI_COLS    = ["Ticker","Price","CurrVol","AvgVol","RSI9","Pattern"]
RSI_RENAME  = {"CurrVol":"Curr Vol","AvgVol":"Avg Vol","RSI9":"RSI(9)"}
GL_COLS     = ["Ticker","Price","AvgVol","EMA250","PctAboveEMA","RS","Stoch"]
GL_RENAME   = {"AvgVol":"Avg Vol","EMA250":"EMA(250)","PctAboveEMA":"% > EMA","RS":"IBD-RS","Stoch":"Stoch(5,1)"}
SP_COLS     = ["Ticker","Price","Change","Volume","Side"]


# ══════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════

with st.sidebar:
    st.title("📈 US Market Screener")
    st.markdown("---")

    screen_choice = st.radio(
        "Run Screen",
        ["RSI Screen", "Green Line", "S&P 500 Movers"],
        label_visibility="collapsed",
    )
    st.markdown("---")

    if screen_choice == "RSI Screen":
        st.subheader("⚙️ RSI Filters")
        rsi_min_price  = st.number_input("Min Price ($)",  value=90.0,  step=5.0,  format="%.0f")
        rsi_avg_vol_k  = st.number_input("Avg Vol (k)",    value=500,   step=100)
        rsi_curr_vol_k = st.number_input("Curr Vol (k)",   value=1000,  step=100)
        rsi9_ob        = st.number_input("RSI(9) OB  >",  value=80.0,  step=1.0,  format="%.0f")
        rsi9_os        = st.number_input("RSI(9) OS  <",  value=20.0,  step=1.0,  format="%.0f")
        st.markdown("---")
        scan_btn = st.button("⟳  SCAN RSI", type="primary", use_container_width=True)
        if st.session_state.rsi_done:
            st.caption(f"Last: {len(st.session_state.rsi_ob)} OB · {len(st.session_state.rsi_os)} OS")

    elif screen_choice == "Green Line":
        st.subheader("⚙️ Green Line Filters")
        gl_min_price = st.number_input("Min Price ($)",  value=10.0,  step=5.0,  format="%.0f")
        gl_avg_vol_k = st.number_input("Avg Vol (k)",    value=200,   step=100)
        gl_rs_thr    = st.number_input("IBD-RS >",       value=80.0,  step=1.0,  format="%.0f")
        st.caption(
            "Rules:\n"
            "1. Close > EMA(250)\n"
            "2. IBD-RS percentile > threshold\n"
            "3. Stoch(5,1) < 20 and turning up\n\n"
            "⚠️ Needs 15mo data — runs slower than RSI screen."
        )
        st.markdown("---")
        scan_btn = st.button("⟳  SCAN GREEN LINE", type="primary", use_container_width=True)
        if st.session_state.gl_done:
            st.caption(f"Last: {len(st.session_state.gl_results)} results")

    else:  # S&P Movers
        st.subheader("⚙️ S&P 500 Movers")
        st.caption(
            "Top 3 gainers and top 3 losers\n"
            "by daily % change.\n\n"
            "Universe: S&P 500 components\n"
            "Source: Wikipedia + yfinance"
        )
        st.markdown("---")
        scan_btn = st.button("⟳  FETCH MOVERS", type="primary", use_container_width=True)

    st.markdown("---")
    st.caption("NASDAQ · NYSE · AMEX\nClick row to chart · Double-click → Yahoo")


# ══════════════════════════════════════════════════════════════════
#  TABS
# ══════════════════════════════════════════════════════════════════

tab_rsi, tab_gl, tab_sp, tab_calc, tab_guide = st.tabs([
    "📊  RSI Screen",
    "🟢  Green Line",
    "🏆  S&P Movers",
    "🧮  Calculator",
    "📖  Pattern Guide",
])


# ══════════════════════════════════════════════════════════════════
#  TAB 1 — RSI SCREEN
# ══════════════════════════════════════════════════════════════════

with tab_rsi:
    if scan_btn and screen_choice == "RSI Screen":
        avg_vol  = int(rsi_avg_vol_k  * 1_000)
        curr_vol = int(rsi_curr_vol_k * 1_000)
        prog = st.progress(0, text="Fetching ticker universe…")

        def _rsi_cb(pct, msg):
            prog.progress(pct / 100, text=msg)

        try:
            tickers = fetch_us_universe(rsi_min_price)
            prog.progress(0.05, text=f"{len(tickers)} candidates → downloading…")
            ob, os_, cache = screen_rsi(
                tickers, rsi_min_price, avg_vol, curr_vol,
                rsi9_ob, rsi9_os, progress_cb=_rsi_cb,
            )
            st.session_state.rsi_ob    = ob
            st.session_state.rsi_os    = os_
            st.session_state.rsi_cache = cache
            st.session_state.rsi_done  = True
            prog.empty()
        except Exception as e:
            prog.empty(); st.error(f"Scan failed: {e}")

    if st.session_state.rsi_done:
        ob  = st.session_state.rsi_ob
        os_ = st.session_state.rsi_os

        if ob or os_:
            all_rows = (
                [{**{c: r[c] for c in RSI_COLS}, "List":"Overbought"} for r in ob] +
                [{**{c: r[c] for c in RSI_COLS}, "List":"Oversold"}   for r in os_]
            )
            csv_buf = io.StringIO()
            pd.DataFrame(all_rows).to_csv(csv_buf, index=False)
            st.download_button("⬇  Export CSV", data=csv_buf.getvalue(),
                               file_name=f"rsi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                               mime="text/csv")

        ob_tab, os_tab = st.tabs([
            f"🔴  Overbought  ({len(ob)})",
            f"🟢  Oversold    ({len(os_)})",
        ])

        def _rsi_tab(rows, key):
            if not rows:
                st.info("No results matching current filters.")
                return
            df = pd.DataFrame(rows)[RSI_COLS].rename(columns=RSI_RENAME)
            ev = st.dataframe(df, use_container_width=True, hide_index=True,
                              selection_mode="single-row", on_select="rerun",
                              height=min(400, 36*len(rows)+38), key=f"rsi_{key}")
            if ev.selection.rows:
                idx     = ev.selection.rows[0]
                ticker  = rows[idx]["Ticker"]
                pattern = rows[idx]["Pattern"]
                hist    = st.session_state.rsi_cache.get(ticker)
                if hist is not None:
                    st.plotly_chart(make_chart(ticker, hist, pattern, rsi9_ob, rsi9_os),
                                    use_container_width=True)
                    if pattern != "—":
                        for p in [x.strip() for x in pattern.split(",")]:
                            entry = PATTERN_DESCRIPTIONS.get(p)
                            if entry:
                                with st.expander(f"▸ {p}  —  {entry[0]}"):
                                    st.write(entry[1])

        with ob_tab: _rsi_tab(ob, "ob")
        with os_tab: _rsi_tab(os_, "os")
    else:
        st.info("Select **RSI Screen** in the sidebar and click **⟳ SCAN RSI**.")


# ══════════════════════════════════════════════════════════════════
#  TAB 2 — GREEN LINE
# ══════════════════════════════════════════════════════════════════

with tab_gl:
    if scan_btn and screen_choice == "Green Line":
        avg_vol = int(gl_avg_vol_k * 1_000)
        prog = st.progress(0, text="Fetching ticker universe…")

        def _gl_cb(pct, msg):
            prog.progress(pct / 100, text=msg)

        try:
            tickers = fetch_us_universe(gl_min_price)
            prog.progress(0.03, text=f"{len(tickers)} candidates → downloading 15mo data…")
            results, cache = screen_green_line(
                tickers, gl_min_price, avg_vol, gl_rs_thr,
                progress_cb=_gl_cb,
            )
            st.session_state.gl_results = results
            st.session_state.gl_cache   = cache
            st.session_state.gl_done    = True
            prog.empty()
        except Exception as e:
            prog.empty(); st.error(f"Scan failed: {e}")

    if st.session_state.gl_done:
        results = st.session_state.gl_results
        st.caption(
            "⚠️ IBD-RS is approximated as 12-month price return percentile within "
            "this scan universe — not identical to SCTR. Expect some divergence from "
            "StockCharts rankings."
        )

        if results:
            csv_buf = io.StringIO()
            pd.DataFrame(results)[GL_COLS].to_csv(csv_buf, index=False)
            st.download_button("⬇  Export CSV", data=csv_buf.getvalue(),
                               file_name=f"greenline_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                               mime="text/csv")

            df = pd.DataFrame(results)[GL_COLS].rename(columns=GL_RENAME)
            ev = st.dataframe(df, use_container_width=True, hide_index=True,
                              selection_mode="single-row", on_select="rerun",
                              height=min(500, 36*len(results)+38), key="gl_tbl")
            if ev.selection.rows:
                idx    = ev.selection.rows[0]
                ticker = results[idx]["Ticker"]
                hist   = st.session_state.gl_cache.get(ticker)
                if hist is not None:
                    st.plotly_chart(make_gl_chart(ticker, hist), use_container_width=True)
        else:
            st.info(
                "No stocks matched all three Green Line rules simultaneously.\n\n"
                "This is normal — the setup is rare by design. Try lowering the "
                "IBD-RS threshold or min price filter."
            )
    else:
        st.info("Select **Green Line** in the sidebar and click **⟳ SCAN GREEN LINE**.")


# ══════════════════════════════════════════════════════════════════
#  TAB 3 — S&P 500 MOVERS
# ══════════════════════════════════════════════════════════════════

with tab_sp:
    if scan_btn and screen_choice == "S&P 500 Movers":
        prog = st.progress(0, text="Fetching S&P 500 components…")

        def _sp_cb(pct, msg):
            prog.progress(pct / 100, text=msg)

        try:
            results = screen_sp500_movers(progress_cb=_sp_cb)
            st.session_state.sp_results = results
            st.session_state.sp_done    = True
            prog.empty()
        except Exception as e:
            prog.empty(); st.error(f"Failed: {e}")

    if st.session_state.sp_done:
        results  = st.session_state.sp_results
        gainers  = [r for r in results if r["Side"] == "Gainer"]
        losers   = [r for r in results if r["Side"] == "Loser"]

        g_col, l_col = st.columns(2)

        with g_col:
            st.markdown("### 🟢 Top Gainers")
            for r in gainers:
                st.markdown(
                    f"<div class='metric-box metric-green'>"
                    f"<b style='font-size:1.1rem'>{r['Ticker']}</b>&nbsp;&nbsp;"
                    f"<span style='color:#3fb950;font-size:1.2rem;font-weight:700'>{r['Change']}</span><br>"
                    f"<span style='color:#6e7f96;font-size:0.85rem'>"
                    f"Price {r['Price']} &nbsp;·&nbsp; Vol {r['Volume']}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
                if st.button(f"Chart {r['Ticker']}", key=f"gc_{r['Ticker']}"):
                    import yfinance as yf
                    hist = yf.download(r["Ticker"], period="4mo", interval="1d",
                                       auto_adjust=True, progress=False)
                    if isinstance(hist.columns, pd.MultiIndex):
                        hist.columns = hist.columns.droplevel(1)
                    if not hist.empty:
                        st.plotly_chart(make_chart(r["Ticker"], hist, "—"),
                                        use_container_width=True)

        with l_col:
            st.markdown("### 🔴 Top Losers")
            for r in losers:
                st.markdown(
                    f"<div class='metric-box metric-red'>"
                    f"<b style='font-size:1.1rem'>{r['Ticker']}</b>&nbsp;&nbsp;"
                    f"<span style='color:#f85149;font-size:1.2rem;font-weight:700'>{r['Change']}</span><br>"
                    f"<span style='color:#6e7f96;font-size:0.85rem'>"
                    f"Price {r['Price']} &nbsp;·&nbsp; Vol {r['Volume']}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
                if st.button(f"Chart {r['Ticker']}", key=f"lc_{r['Ticker']}"):
                    import yfinance as yf
                    hist = yf.download(r["Ticker"], period="4mo", interval="1d",
                                       auto_adjust=True, progress=False)
                    if isinstance(hist.columns, pd.MultiIndex):
                        hist.columns = hist.columns.droplevel(1)
                    if not hist.empty:
                        st.plotly_chart(make_chart(r["Ticker"], hist, "—"),
                                        use_container_width=True)
    else:
        st.info("Select **S&P 500 Movers** in the sidebar and click **⟳ FETCH MOVERS**.")


# ══════════════════════════════════════════════════════════════════
#  TAB 4 — CALCULATOR
# ══════════════════════════════════════════════════════════════════

with tab_calc:
    st.subheader("Stock Calculator — RSI & Std Dev / Z-Score")
    st.caption("Downloads OHLC data and builds a formatted Excel workbook.")

    if not HAS_XL:
        st.error("openpyxl is not installed. Add it to requirements.txt.")
    else:
        src_choice    = st.radio("Data Source",
                                 ["Yahoo Finance  (stocks, ETFs)",
                                  "Twelve Data  (FX & spot metals)"],
                                 horizontal=True)
        is_twelvedata = src_choice.startswith("Twelve Data")

        td_api_key = ""
        if is_twelvedata:
            secret_key  = st.secrets.get("TWELVEDATA_API_KEY", "")
            default_key = st.session_state.get("td_api_key", secret_key)
            td_api_key  = st.text_input(
                "Twelve Data API Key", value=default_key, type="password",
                help="Stored in Secrets by default — override here for this session only.",
            )
            st.session_state["td_api_key"] = td_api_key
            if secret_key and td_api_key == secret_key:
                st.caption("🔑 Using key from Streamlit Secrets")
            st.caption(
                "Examples: **XAU/USD** (gold spot) · **XAG/USD** (silver) · "
                "**EUR/USD** · **GBP/JPY**"
            )

        c1, c2 = st.columns(2)
        with c1:
            calc_mode   = st.selectbox("Calculator", ["RSI", "Std Dev / Z-Score"])
            default_tk  = "XAU/USD" if is_twelvedata else "AAPL"
            calc_ticker = st.text_input("Ticker Symbol", value=default_tk).upper().strip()
            calc_period = st.number_input(
                "RSI Period" if calc_mode == "RSI" else "Rolling Window",
                value=9 if calc_mode == "RSI" else 20, min_value=2,
            )
        with c2:
            calc_interval = st.selectbox("Data Interval",
                                         ["Daily","Weekly","Monthly","Annually"])
            today = datetime.today()
            calc_start = st.date_input("Start Date",
                                       value=today - timedelta(days=730),
                                       min_value=datetime(1900, 1, 1),
                                       max_value=today, format="YYYY-MM-DD",
                                       help="Type any year e.g. 1990-01-01")
            calc_end   = st.date_input("End Date", value=today,
                                       min_value=datetime(1900, 1, 1),
                                       max_value=today, format="YYYY-MM-DD")

        if st.button("⬇  Download & Build Excel", type="primary"):
            if not calc_ticker:
                st.warning("Enter a ticker symbol.")
            elif is_twelvedata and not td_api_key:
                st.warning("Enter your Twelve Data API key first.")
            else:
                with st.spinner(f"Downloading {calc_ticker} and building Excel…"):
                    try:
                        iv_map = {"Daily":"1d","Weekly":"1wk","Monthly":"1mo","Annually":"1y"}
                        sf_map = {"Daily":"","Weekly":"_Weekly","Monthly":"_Monthly","Annually":"_Annual"}
                        df_ohlc = download_ohlc(
                            calc_ticker, str(calc_start), str(calc_end),
                            iv_map[calc_interval],
                            source="twelvedata" if is_twelvedata else "yahoo",
                            api_key=td_api_key,
                        )
                        buf = io.BytesIO()
                        safe_ticker = calc_ticker.replace("/", "-")
                        if calc_mode == "RSI":
                            build_rsi_excel(df_ohlc, calc_ticker, calc_period, buf)
                            fname = f"{safe_ticker}_RSI{sf_map[calc_interval]}.xlsx"
                        else:
                            build_sd_excel(df_ohlc, calc_ticker, calc_period, buf)
                            fname = f"{safe_ticker}_SD_ZScore{sf_map[calc_interval]}.xlsx"
                        buf.seek(0)
                        st.success(f"✓  Built {len(df_ohlc)} rows of {calc_interval.lower()} data")
                        st.download_button(
                            label=f"📂  Save {fname}", data=buf, file_name=fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    except Exception as e:
                        st.error(f"Error: {e}")


# ══════════════════════════════════════════════════════════════════
#  TAB 5 — PATTERN GUIDE
# ══════════════════════════════════════════════════════════════════

with tab_guide:
    st.subheader("Chart Pattern Reference")
    st.caption("All patterns are heuristic — treat as a first-pass filter, not a signal.")
    st.markdown("---")
    for pattern, (label, body) in PATTERN_DESCRIPTIONS.items():
        with st.expander(f"▸  {pattern}  —  {label}"):
            st.write(body)
